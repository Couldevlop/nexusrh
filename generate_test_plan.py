"""
NexusRH CI — Cahier de recettes professionnel
Génère: NEXUSRH_CI_Test_Plan.xlsx
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ── Palette ───────────────────────────────────────────────────────────────────
C_NAVY    = "1E3A5F"   # header principal
C_ORANGE  = "E85D04"   # accent OpenLab
C_ORANGE2 = "F48C06"   # accent secondaire
C_BGLIGHT = "F4F6FB"   # fond alterné
C_WHITE   = "FFFFFF"
C_GREY    = "DDE3EF"
C_GREEN   = "16A34A"
C_RED     = "DC2626"
C_AMBER   = "D97706"
C_BLUE    = "2563EB"
C_SLATE   = "64748B"
C_PASS    = "DCFCE7"   # fond passé
C_FAIL    = "FEE2E2"   # fond échoué
C_BLOCK   = "FEF9C3"   # fond bloqué
C_RUN     = "DBEAFE"   # fond en cours

STATUS_LIST = '"✅ Passé,❌ Échoué,⚠️ Bloqué,▶️ En cours,⬜ Non exécuté"'
PRIO_LIST   = '"P1 — Critique,P2 — Haute,P3 — Moyenne,P4 — Basse"'
ENV_LIST    = '"Production,Preprod,Local"'

TODAY       = date.today().strftime("%d/%m/%Y")
VERSION     = "v1.0"

def fill(hex_color, fill_type="solid"):
    return PatternFill(fill_type=fill_type, fgColor=hex_color)

def font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="C0C8D8")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="1E3A5F")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value="", bold=False, fgColor=None, fontColor="000000",
             halign="left", wrap=False, size=10, border=True, valign="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fontColor, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if fgColor:
        c.fill = PatternFill(fill_type="solid", fgColor=fgColor)
    if border:
        c.border = thin_border()
    return c

# ── Données de test ───────────────────────────────────────────────────────────
MODULES = {
    "AUTH": {
        "label": "Authentification & Sécurité",
        "color": "1E3A5F",
        "cases": [
            ("P1","Login super_admin valide","Connexion avec superadmin@nexusrh-ci.com / SuperAdmin1234!",
             "API déployée, DB seedée","1. Ouvrir nexusrh.openlabconsulting.com\n2. Saisir email + mdp\n3. Cliquer Connexion",
             "Token JWT reçu, redirection vers /platform/dashboard, sidebar plateforme affichée"),
            ("P1","Login admin tenant valide","Connexion admin@sotra.ci / Admin1234!",
             "Tenant SOTRA seedé","1. Saisir admin@sotra.ci\n2. Saisir Admin1234!\n3. Cliquer Connexion",
             "Redirection /dashboard, thème orange SOTRA appliqué, sidebar RH complète"),
            ("P1","Login mot de passe incorrect → message d'erreur visible","Vérifier que le message 401 remonte au formulaire (plus de rechargement silencieux)",
             "Correctif intercepteur 401 déployé","1. Saisir un email valide\n2. Saisir un mauvais mot de passe\n3. Soumettre",
             "Message 'Email ou mot de passe incorrect' affiché dans le formulaire, pas de rechargement de page"),
            ("P2","Login email inexistant","Email qui n'existe dans aucun tenant",
             "","1. Saisir email@inconnu.com\n2. Mdp quelconque\n3. Soumettre",
             "Message d'erreur 401 affiché sans plantage"),
            ("P1","Redirection rôle employee → /mon-espace","employe@sotra.ci doit atterrir sur son espace",
             "Seed SOTRA OK","1. Connecter employe@sotra.ci\n2. Observar la redirection",
             "Redirection automatique vers /mon-espace, sidebar réduite (5 items seulement)"),
            ("P1","Redirection rôle manager → /dashboard","manager@sotra.ci → /dashboard version manager",
             "","1. Connecter manager@sotra.ci\n2. Observer dashboard",
             "Dashboard manager affiché (équipe, demandes à valider), pas les KPIs admin"),
            ("P1","Redirection super_admin → /platform/dashboard","Aucun accès aux données RH",
             "","1. Connecter superadmin@nexusrh-ci.com\n2. Tenter /dashboard",
             "Redirigé vers /platform/dashboard, sidebar plateforme uniquement"),
            ("P2","Refresh automatique token JWT expiré","Le refresh token permet de renouveler le JWT sans déconnexion",
             "Token expiré simulé","1. Connecter\n2. Attendre expiration JWT ou modifier exp dans DevTools\n3. Effectuer une requête API",
             "Token rafraîchi silencieusement, requête relancée, utilisateur non déconnecté"),
            ("P2","Déconnexion propre","Store Zustand vidé, redirection /login",
             "Session active","1. Cliquer sur Déconnexion\n2. Vérifier localStorage/sessionStorage",
             "Token supprimé, store authStore vidé, redirection /login, route protégée inaccessible"),
            ("P1","Route protégée sans token → /login","Un utilisateur non connecté tente d'accéder /dashboard",
             "","1. Sans session, accéder directement à /dashboard","Redirection automatique vers /login"),
            ("P2","must_change_password = false après seed","Les comptes seedés n'exigent pas de changement de mdp",
             "Seed v2 avec last_login_at déployé","1. Connecter n'importe quel compte seedé",
             "Connexion directe au dashboard, pas de popup changement de mot de passe"),
            ("P3","Activation MFA (QR code TOTP)","L'utilisateur active la double authentification",
             "","1. Mon profil → Sécurité → Activer MFA\n2. Scanner QR code avec app TOTP\n3. Valider avec code",
             "MFA activé, codes de secours affichés, connexion suivante demande OTP"),
            ("P2","MFA code OTP valide","Connexion avec MFA activé",
             "MFA activé sur le compte","1. Connecter\n2. Saisir code TOTP valide",
             "Accès accordé, token JWT émis"),
            ("P2","MFA code OTP invalide → rejet","",
             "MFA activé","1. Connecter\n2. Saisir code TOTP périmé ou incorrect",
             "Erreur 401, message explicite, pas de token émis"),
            ("P1","Isolation multi-tenant","Token SOTRA ne donne pas accès aux données Cabinet Expertise CI",
             "2 tenants seedés","1. Connecter admin@sotra.ci\n2. Inspecter les requêtes API\n3. Tenter d'appeler /api/employees avec schéma cabinet",
             "403 Forbidden ou données vides, jamais de fuite cross-tenant"),
            ("P1","[NON-REGRESSION] Seed DO UPDATE password_hash",
             "Régression résolue 2026-05-15 : seed ON CONFLICT DO NOTHING ne rafraîchissait pas les hashes SOTRA/Cabinet (login 401 sur tous les comptes)",
             "Re-seed exécuté en prod","1. kubectl apply seed-job-prod.yaml ou pnpm run db:seed\n2. Tenter login admin@sotra.ci / Admin1234!",
             "Login OK 200 — le seed met à jour password_hash si l'utilisateur existe déjà"),
            ("P1","[NON-REGRESSION] Script reset-admin-passwords.sql",
             "Récupération prod sans rebuild — applique 8 UPDATE bcrypt 12 rounds",
             "Accès kubectl prod","1. kubectl exec -i postgres-0 -- psql -U nexusrh -d nexusrh < scripts/reset-admin-passwords.sql\n2. Vérifier les is_active=true\n3. Login admin@sotra.ci",
             "8 lignes UPDATE OK, login fonctionnel sur les 8 comptes documentés"),
            ("P2","[NON-REGRESSION] Script reset-admin-passwords.ts standalone",
             "pnpm run admin:reset-passwords:dry-run puis admin:reset-passwords",
             "DB locale ou prod accessible","1. pnpm --filter @nexusrhci/api run admin:reset-passwords:dry-run\n2. Vérifier la liste des comptes ciblés\n3. Lancer admin:reset-passwords",
             "Mode dry-run liste les 8 cibles. Lancement réel : ✓ pour chaque compte. Bcrypt 12 rounds, OWASP A02 respecté"),
        ]
    },
    "PLATFORM": {
        "label": "Portail Super Admin",
        "color": "312E81",
        "cases": [
            ("P1","KPIs plateforme corrects","Dashboard /platform/dashboard affiche tenants actifs, trials, MRR",
             "2+ tenants seedés","1. Connecter superadmin\n2. Observer KPI cards",
             "Valeurs cohérentes avec la base (ex: 3 tenants actifs = 3 affiché)"),
            ("P1","Création tenant — wizard 3 étapes","Créer un nouveau tenant via le formulaire",
             "Super admin connecté","1. /platform/tenants/new\n2. Étape 1: nom, slug, plan, secteur, ville\n3. Étape 2: admin email/prénom/nom\n4. Étape 3: couleurs",
             "Tenant créé, schema PostgreSQL provisionné, email admin envoyé, tempPassword affiché"),
            ("P2","Slug auto-généré depuis le nom","Le slug se pré-remplit en minuscules-sans-espaces",
             "","1. Saisir nom entreprise 'Tech Corp'\n2. Observer le champ slug",
             "Slug = 'tech-corp' (minuscules, tirets, sans caractères spéciaux)"),
            ("P2","Slug déjà existant → erreur","",
             "Tenant 'sotra' existant","1. Créer un tenant avec slug 'sotra'",
             "Erreur explicite 'Slug déjà utilisé', formulaire non soumis"),
            ("P2","Seed données démo (seedDemoData)","Cocher l'option inject démo lors création",
             "","1. Wizard création\n2. Cocher 'Injecter données de démonstration'\n3. Créer",
             "8 employés, 3 mois de bulletins créés dans le nouveau tenant, non bloquant"),
            ("P1","tempPassword retourné dans réponse","En cas d'échec email, le mdp reste accessible",
             "","1. Créer un tenant\n2. Observer la réponse écran de confirmation",
             "Mot de passe temporaire affiché dans l'encart de succès"),
            ("P1","Liste tenants paginée","Tableau /platform/tenants",
             "3+ tenants","1. Accéder /platform/tenants",
             "Tableau avec nom, plan, ville, statut, nb users, nb employés, actions"),
            ("P2","Filtrage tenants par statut","Filtre active/trial/suspended",
             "","1. Utiliser le filtre statut","Tableau filtré correctement"),
            ("P1","Détail tenant — 4 onglets","Apparence, Plan, Modules, Données",
             "","1. Cliquer sur un tenant dans la liste",
             "4 onglets affichés et fonctionnels"),
            ("P2","Modification apparence — couleur primaire","Changer couleur et observer preview live",
             "","1. Onglet Apparence\n2. Changer couleur primaire\n3. Observer prévisualisation\n4. Sauvegarder",
             "Couleur mise à jour, preview reflète le changement immédiatement"),
            ("P2","Modification plan tenant","Changer de trial → business",
             "","1. Onglet Plan\n2. Sélectionner Business\n3. Sauvegarder",
             "Plan mis à jour dans platform.tenants, maxUsers/maxEmployees ajustés selon PLAN_DEFAULTS"),
            ("P2","Suspension tenant","Tenant passé en statut suspended",
             "","1. Onglet Plan → Suspendre",
             "Status = suspended, login des users du tenant retourne erreur"),
            ("P2","Réactivation tenant","",
             "Tenant suspendu","1. Onglet Plan → Réactiver",
             "Status = active, users peuvent se reconnecter"),
            ("P2","Consulter utilisateurs tenant","Lecture seule des users d'un tenant",
             "","1. /platform/tenants/:id/users",
             "Liste avec email, rôle, statut, date création — pas de création depuis ici"),
            ("P2","Suspension utilisateur tenant","",
             "","1. Trouver un utilisateur\n2. Cliquer Suspendre",
             "is_active = false, cet user ne peut plus se connecter"),
            ("P2","Reset admin tenant","Regénérer un mdp temporaire pour l'admin",
             "","1. POST /platform/tenants/:id/reset-admin",
             "Nouveau tempPassword retourné + email envoyé"),
            ("P3","Diagnostic admin-status","GET /platform/tenants/:id/admin-status",
             "","1. Appeler l'endpoint via Swagger",
             "JSON avec schemaExists, adminUser.isActive, hasPasswordHash, issue"),
            ("P1","Logs activité cross-tenant","/platform/logs",
             "","1. Accéder /platform/logs",
             "Logs triés par date, avec tenant, user, action, IP"),
            ("P3","Alerte trials expirant < 7 jours","Dashboard plateforme signale les essais proches de l'expiration",
             "Tenant trial.endsAt dans < 7j","1. Observer le panneau alertes du dashboard",
             "Badge rouge avec le nom du tenant et date d'expiration"),
            ("P2","Thématisation dynamique au login","Au login, les couleurs tenant s'appliquent",
             "","1. Accéder page login SOTRA\n2. Saisir admin@sotra.ci",
             "Bouton login orange (#E85D04), logo SOTRA visible avant même de se connecter"),
            ("P3","Graphique croissance tenants 12 mois","LineChart sur le dashboard plateforme",
             "","1. Observer le graphique dashboard plateforme",
             "Courbe affichée avec données des 12 derniers mois"),
            ("P2","MRR estimé FCFA","KPI card affiche le revenu mensuel récurrent estimé",
             "","1. Observer KPI dashboard plateforme",
             "Valeur en FCFA calculée selon plan × tarif indicatif"),
            ("P2","Feature flags par tenant","Activer/désactiver modules individuellement",
             "","1. Onglet Modules d'un tenant\n2. Désactiver 'IA Assistant'\n3. Connecter admin du tenant",
             "Module IA n'apparaît pas dans la sidebar du tenant"),
        ]
    },
    "EMPLOYEES": {
        "label": "Gestion des Employés",
        "color": "065F46",
        "cases": [
            ("P1","Création employé — champs obligatoires","Prénom, nom, email, département, poste, salaire brut FCFA",
             "Admin connecté, département créé","1. /employees/new\n2. Remplir tous les champs requis\n3. Sauvegarder",
             "Employé créé, visible dans la liste, aucune décimale dans le salaire"),
            ("P1","Email employé unique dans le tenant","",
             "","1. Créer 2 employés avec le même email",
             "Erreur 409, message 'Email déjà utilisé dans ce tenant'"),
            ("P2","Validation NNI — format ivoirien","Format CI + 9 chiffres",
             "","1. Saisir NNI 'CI123456789'\n2. Saisir NNI invalide 'AB12'",
             "NNI valide accepté, NNI invalide rejeté avec message d'erreur"),
            ("P2","Validation CNPS — format CI","Format CI + 8 chiffres + lettre",
             "","1. Saisir CI12345678A (valide)\n2. Saisir 12345 (invalide)",
             "Valide accepté, invalide rejeté"),
            ("P2","Validation téléphone Mobile Money CI","+225 07/05 + 8 chiffres",
             "","1. Saisir +225 07 12 34 56 78 (valide)\n2. Saisir +33 6 12 (invalide)",
             "Format CI accepté, autre rejeté"),
            ("P1","Liste employés — pagination","Affichage par page avec navigation",
             "80+ employés SOTRA","1. Accéder /employees",
             "Pagination fonctionnelle, 20-50 par page, nombre total affiché"),
            ("P1","Recherche employé par nom","Champ de recherche filtrant en temps réel",
             "","1. Taper 'Kouassi' dans la recherche",
             "Liste filtrée aux employés correspondants"),
            ("P2","Filtrage par département","Sélecteur département",
             "","1. Sélectionner département 'Exploitation'\n2. Observer la liste",
             "Seuls les employés du département Exploitation affichés"),
            ("P1","Fiche employé — consultation","Tous les champs visibles",
             "","1. Cliquer sur un employé dans la liste",
             "Fiche complète: identité, poste, salaire FCFA, NNI, CNPS, Mobile Money, ancienneté calculée"),
            ("P1","Modification employé","Mise à jour des données",
             "","1. Ouvrir fiche\n2. Modifier salaire\n3. Sauvegarder",
             "Données mises à jour, historique hr_events enregistré"),
            ("P2","Désactivation employé","is_active = false, employé absent des listings actifs",
             "","1. Fiche employé → Désactiver",
             "Employé marqué inactif, n'apparaît plus dans liste active"),
            ("P2","Calcul ancienneté affiché","Années et mois depuis hire_date",
             "","1. Ouvrir fiche employé embauché il y a 3 ans 2 mois",
             "Affiche '3 ans 2 mois' correctement"),
            ("P3","Import CSV employés","Upload d'un fichier CSV pour création en masse",
             "Fichier CSV valide préparé","1. /employees/import\n2. Uploader CSV\n3. Valider",
             "Employés créés en masse, rapport d'import (succès/erreurs)"),
            ("P3","Export liste employés","Téléchargement CSV/Excel",
             "","1. Bouton Export dans la liste\n2. Choisir format",
             "Fichier téléchargé avec tous les champs, encodage UTF-8, FCFA entiers"),
            ("P3","Upload photo profil","Image recadrée et sauvegardée",
             "","1. Fiche employé → Photo\n2. Uploader une image\n3. Recadrer\n4. Sauvegarder",
             "Photo visible dans la fiche et dans le header"),
            ("P1","Manager ne voit que son équipe","Filtrage automatique par department_id du manager",
             "manager@sotra.ci lié à département Exploitation","1. Connecter manager@sotra.ci\n2. Accéder /employees",
             "Seuls les employés de son département listés"),
            ("P1","Employee ne peut pas accéder /employees","Redirection /mon-espace",
             "","1. Connecter employe@sotra.ci\n2. Tenter /employees",
             "Redirection automatique vers /mon-espace"),
            ("P2","Mobile Money affiché sur fiche","Provider + numéro formaté",
             "","1. Ouvrir fiche d'un employé avec Wave enregistré",
             "Affiche 'Wave — +225 07 XX XX XX XX'"),
        ]
    },
    "PAYROLL": {
        "label": "Moteur de Paie CI (CNPS + ITS)",
        "color": "7C3AED",
        "cases": [
            ("P1","CNPS Retraite salarié 6.3% — plafond 1 647 315 FCFA","Vérifier le calcul exact pour un salaire de 200 000 FCFA",
             "Moteur CI déployé","1. Créer bulletin pour salaire 200 000 FCFA\n2. Vérifier rubrique CNPS retraite sal",
             "200 000 × 6.3% = 12 600 FCFA (< plafond → pas d'écrêtement)"),
            ("P1","CNPS Retraite salarié — plafond appliqué","Salaire > 1 647 315 FCFA",
             "","1. Créer bulletin pour salaire 2 000 000 FCFA",
             "CNPS retraite sal = 1 647 315 × 6.3% = 103 781 FCFA (pas 2 000 000 × 6.3%)"),
            ("P1","CNPS AT commerce 2% — plafond 70 000 FCFA","Tenant secteur Commerce",
             "Tenant services (atRate=2%)","1. Générer bulletin pour salaire 200 000 FCFA\n2. Vérifier rubrique AT",
             "AT = min(200 000, 70 000) × 2% = 1 400 FCFA"),
            ("P1","CNPS AT BTP 3% — plafond 70 000 FCFA","Tenant secteur Transport (SOTRA atRate=3%)",
             "Tenant SOTRA connecté","1. Générer bulletin SOTRA, salaire 180 000 FCFA",
             "AT = min(180 000, 70 000) × 3% = 2 100 FCFA"),
            ("P1","CNPS Prestations Familiales 5%","Part patronale seulement, plafond 70 000",
             "","1. Vérifier rubrique PF sur bulletin",
             "PF = min(BRUT, 70 000) × 5%, salarié ne paie rien sur cette branche"),
            ("P1","CNPS Maternité 0.75% — patronal","",
             "","1. Vérifier rubrique maternité",
             "Maternité = min(BRUT, 70 000) × 0.75%"),
            ("P1","ITS barème 0% — salaire net imposable ≤ 75 000 FCFA","Salaire SMIG = 60 000 FCFA",
             "","1. Créer bulletin SMIG = 60 000 FCFA\n2. Calculer net imposable = 60 000 × 0.85 - CNPS",
             "ITS = 0 FCFA"),
            ("P1","ITS barème 1.5% — tranche 75 001 à 240 000 FCFA","Salaire net imposable dans cette tranche",
             "","1. Salaire brut 120 000 FCFA\n2. Net imposable = 120 000×0.85 - CNPS ≈ 94 440",
             "ITS = (94 440 - 75 000) × 1.5% ≈ 292 FCFA"),
            ("P1","ITS barème 5% — tranche 240 001 à 800 000 FCFA","",
             "","1. Salaire brut 400 000 FCFA","ITS calculé selon barème progressif, taux 5% sur la tranche"),
            ("P1","ITS barème 10% — tranche 800 001 à 2 000 000 FCFA","",
             "","1. Salaire brut 1 200 000 FCFA","ITS avec taux 10% sur la tranche concernée"),
            ("P1","ITS barème 15% — tranche > 2 000 000 FCFA","",
             "","1. Salaire brut 3 000 000 FCFA","ITS avec taux 15% sur la tranche > 2M"),
            ("P1","Abattement ITS 15% appliqué avant barème","Net imposable = BRUT × 0.85",
             "","1. Salaire 200 000 FCFA\n2. Vérifier base de calcul ITS",
             "Base ITS = 200 000 × 0.85 - CNPS sal, PAS directement 200 000"),
            ("P2","Crédit d'impôt marié sans enfant — 5 500 FCFA/mois","",
             "","1. Employé marié, 0 enfant\n2. Vérifier crédit ITS",
             "ITS final = max(0, ITS brut - 5 500)"),
            ("P2","Crédit d'impôt 1 enfant — 3 000 FCFA supplémentaires","",
             "","1. Employé marié, 1 enfant\n2. Crédit = 5 500 + 3 000 = 8 500 FCFA",
             "ITS final déduit du crédit cumulé"),
            ("P2","Crédit d'impôt 3+ enfants — 9 000 FCFA supplémentaires","",
             "","1. Employé marié, 3 enfants\n2. Crédit = 5 500 + 9 000 = 14 500 FCFA",
             "ITS final correct"),
            ("P1","Net payable ≥ SMIG 60 000 FCFA","Vérification du plancher légal CI",
             "","1. Créer bulletin pour salaire bas (60 000 FCFA)\n2. Vérifier net payable",
             "Net payable ≥ 60 000 FCFA, jamais en dessous du SMIG pour temps plein"),
            ("P2","Prorata jours travaillés","BRUT_PRORATA = BRUT × (jours_travaillés / jours_ouvrables)",
             "","1. Employé ayant travaillé 13j sur 26j ouvrables\n2. Vérifier BRUT_PRORATA",
             "Salaire brut = 50% du brut mensuel contractuel"),
            ("P2","Prime de transport intégrée (variable elements)","",
             "","1. Ajouter prime transport 30 000 FCFA dans variable elements",
             "Bulletin inclut prime, brut augmenté, CNPS recalculé"),
            ("P1","Création période de paie","",
             "Admin connecté","1. Paie → Nouvelle période → Mai 2025\n2. Créer",
             "Période créée avec statut 'open'"),
            ("P1","Génération bulletins pour tous les employés actifs","",
             "Période ouverte","1. Période → Générer bulletins\n2. Confirmer",
             "Un bulletin par employé actif, calculs CNPS+ITS correct, pas de décimales"),
            ("P1","Clôture période de paie","",
             "Bulletins générés","1. Période → Clôturer",
             "Statut = 'closed', bulletins figés, plus de modification possible"),
            ("P1","Téléchargement bulletin PDF — employee","employe@sotra.ci télécharge son bulletin",
             "Bulletins générés","1. /mon-espace/bulletins\n2. Cliquer Télécharger sur un bulletin",
             "PDF téléchargé avec logo SOTRA, mentions CNPS, ITS, NNI, net FCFA, pas de décimales"),
            ("P2","Badge 'Nouveau' bulletin non consulté","",
             "Bulletin non encore ouvert par l'employé","1. Connecter employe\n2. Observer /mon-espace/bulletins",
             "Badge 'Nouveau' visible sur les bulletins dont viewed_by_employee_at IS NULL"),
            ("P1","Historique 24 bulletins — mon-espace","",
             "","1. /mon-espace/bulletins",
             "Liste des 24 derniers bulletins, triés du plus récent au plus ancien"),
            ("P1","Cumuls annuels corrects sur bulletin","",
             "","1. Ouvrir un bulletin de décembre\n2. Vérifier cumuls YTD",
             "Cumuls = somme de tous les bulletins de l'année en cours"),
            ("P1","FCFA entiers — zéro décimale","",
             "","1. Ouvrir n'importe quel bulletin\n2. Vérifier tous les montants",
             "Aucun montant avec virgule, tout en FCFA entiers"),
            ("P2","Vue admin — tous les bulletins","Admin peut accéder aux bulletins de tous les employés",
             "","1. Connecter admin\n2. Accéder /payroll/payslips",
             "Grille complète avec filtres par période et employé"),
            ("P2","Employeur cost calculé","BRUT + cotisations patronales",
             "","1. Vérifier le coût employeur affiché sur le bulletin",
             "employer_cost = gross + total_cnps_pat, affiché correctement"),
        ]
    },
    "ABSENCES": {
        "label": "Gestion des Absences",
        "color": "B45309",
        "cases": [
            ("P1","Soldes congés affichés — /mon-espace/absences","2.5 jours ouvrables/mois CI",
             "employe@sotra.ci connecté","1. /mon-espace/absences\n2. Observer les barres de soldes",
             "CP affiché avec acquired, taken, remaining. Barres de progression colorées"),
            ("P1","Nouvelle demande absence — formulaire complet","",
             "","1. Bouton '+ Demander une absence'\n2. Type: CP, dates, motif\n3. Soumettre",
             "Demande créée statut 'pending', notification envoyée au manager"),
            ("P2","Validation dates — début < fin","",
             "","1. Saisir date début > date fin\n2. Soumettre",
             "Erreur Zod côté client + validation serveur, demande non créée"),
            ("P1","Workflow approbation — manager approve","",
             "Demande pending + manager connecté","1. Connecter manager\n2. Voir demandes à valider\n3. Approuver",
             "Statut → 'approved', solde CP décompté, notification employé"),
            ("P2","Workflow — manager refuse","",
             "","1. Manager refuse une demande\n2. Motif de refus saisi",
             "Statut → 'rejected', solde non décompté, notification employé avec motif"),
            ("P2","Annulation demande pending par l'employé","",
             "Demande en statut pending","1. Employé → liste absences → Annuler",
             "Demande supprimée ou statut cancelled, solde inchangé"),
            ("P1","Calcul en jours ouvrables (lundi→samedi CI)","Pas les jours calendaires",
             "","1. Saisir du lundi au vendredi suivant (5j)\n2. Vérifier décompte",
             "5 jours ouvrables décomptés (samedi inclus si travaillé)"),
            ("P2","Jours fériés CI exclus du calcul","Si absence couvre un jour férié CI",
             "Table jours fériés CI 2024/2025 en base","1. Absence couvrant le 07 août 2025 (Fête Nationale)\n2. Vérifier décompte",
             "Jour férié non compté dans les jours d'absence"),
            ("P2","Demi-journée toggle","Absence de demi-journée = 0.5 jour",
             "","1. Cocher 'demi-journée'\n2. Sélectionner date\n3. Soumettre",
             "0.5 jour décompté du solde"),
            ("P1","Historique absences employé — statuts colorés","Badge couleur selon statut",
             "","1. /mon-espace/absences → liste historique",
             "pending=orange, approved=vert, rejected=rouge"),
            ("P1","Admin voit toutes les absences de tous les employés","",
             "Admin connecté","1. /absences (vue admin)",
             "Toutes les absences du tenant, filtrables par employé/type/période"),
            ("P2","Manager voit absences de son équipe seulement","",
             "Manager connecté","1. /absences (vue manager)",
             "Filtrée au département du manager, pas les autres équipes"),
            ("P1","Employee voit ses propres absences seulement","",
             "employe@sotra.ci","1. /mon-espace/absences",
             "Uniquement les absences de l'employé connecté"),
            ("P2","Planification calendrier — absences colorées","",
             "","1. Onglet Calendrier\n2. Observer le mois en cours",
             "Absences apparaissent sur les jours concernés, colorées par type"),
            ("P2","Solde mis à jour après approbation","",
             "","1. Solde CP avant = 15j\n2. Demande 5j approuvée",
             "Solde = 15 - 5 = 10j après approbation"),
            ("P3","Notification manager à la soumission","",
             "SMTP configuré","1. Employé soumet une demande\n2. Vérifier email manager",
             "Email reçu avec détails de la demande et lien d'approbation"),
        ]
    },
    "RECRUTEMENT": {
        "label": "Recrutement & ATS",
        "color": "0E7490",
        "cases": [
            ("P1","Création offre d'emploi","",
             "Admin/RH connecté","1. /recruitment/new\n2. Titre, lieu, type contrat, fourchette salariale FCFA, description\n3. Publier",
             "Offre créée, statut 'open', visible dans le tableau de bord recrutement"),
            ("P2","Clôture offre d'emploi","",
             "Offre ouverte","1. Actions → Clôturer l'offre",
             "Statut = 'closed', n'apparaît plus dans les offres actives"),
            ("P1","Pipeline Kanban — 5 colonnes","new → screening → interview → offer → hired",
             "Offre ouverte avec candidatures","1. Vue Kanban de l'offre",
             "5 colonnes affichées, candidatures déplaçables par drag & drop"),
            ("P1","Ajout candidature","",
             "","1. Offre → Ajouter candidature\n2. Prénom, nom, email, téléphone CI, CV",
             "Candidature créée en stage 'new'"),
            ("P2","Déplacement stage Kanban","",
             "","1. Glisser une candidature de 'screening' vers 'interview'",
             "Stage mis à jour, historique conservé"),
            ("P3","Scoring IA candidature","",
             "ANTHROPIC_API_KEY configuré","1. Ouvrir une candidature\n2. Cliquer 'Analyser avec IA'",
             "Score de compatibilité retourné avec facteurs explicatifs en français"),
            ("P3","Génération contrat OHADA à l'embauche","Stage = hired déclenche la génération du contrat",
             "","1. Déplacer candidature vers 'hired'\n2. Compléter informations contrat",
             "PDF CDI/CDD CI généré avec mentions OHADA obligatoires"),
            ("P2","Fourchette salariale en FCFA","",
             "","1. Créer offre avec salary_min 250 000, salary_max 350 000",
             "Affiché 'De 250 000 à 350 000 FCFA' sans décimales"),
            ("P2","Filtrage offres par statut","",
             "","1. Filtre 'Ouvertes' sur la liste","Seules les offres open affichées"),
            ("P1","[NEW] Onglet Sourcing IA dans Recrutement",
             "Nouvel onglet entre Pipeline et Carrières",
             "API + Web déployés branche feat/multi-pays-recrutement-owasp",
             "1. /recruitment\n2. Cliquer onglet 'Sourcing IA' (icône Sparkles)\n3. Sélectionner une offre 'Chauffeur Bus Senior'",
             "Onglet visible, 3e onglet actif. Affichage automatique des profils sourcés en cache (cards émeraude)"),
            ("P1","[NEW] Visualisation profils sourcés du seed",
             "Le seed pré-remplit 8 profils Chauffeur + 6 profils RH SOTRA + 5 profils OpenLab",
             "DB seedée avec données sourced_profiles",
             "1. Onglet Sourcing IA\n2. Sélectionner offre 'Chauffeur Bus Senior'",
             "8 cards profils visibles avec match_score, ville, disponibilité, plateforme suggérée, salaire FCFA"),
            ("P1","[NEW] Transfert profil unique vers pipeline Kanban",
             "Bouton 'Transférer' sur chaque carte profil",
             "Profils en cache présents","1. Carte profil → bouton 'Transférer'\n2. Observer le badge 'Dans le pipeline'\n3. Aller onglet Pipeline Kanban",
             "Profil créé comme candidature (source=sourced_ai, stage=new) visible dans la colonne Nouveau du Kanban"),
            ("P1","[NEW] Transfert en masse 'Tout transférer'",
             "Bouton 'Tout transférer (N)' transfère tous les profils non transférés",
             "Au moins 3 profils non transférés","1. Bannière émeraude en haut → 'Tout transférer (N)'\n2. Confirmer\n3. Aller Pipeline Kanban",
             "N profils créés en transaction, message succès affiché. Tous présents dans colonne 'Nouveau' du Kanban"),
            ("P1","[NEW] Drag-and-drop Kanban sur profil transféré",
             "Vérifier que les profils issus du sourcing peuvent être déplacés normalement",
             "1+ profil transféré présent","1. Pipeline Kanban\n2. Glisser le profil de 'Nouveau' vers 'Présélection'\n3. Vérifier le PATCH /applications/:id/stage",
             "Stage mis à jour, profil dans la nouvelle colonne, aucune erreur 500"),
            ("P2","[NEW] Mode Compare Claude vs Mistral",
             "Comparaison parallèle des deux modèles IA",
             "ANTHROPIC_API_KEY + MISTRAL_API_KEY configurés","1. Onglet Sourcing IA\n2. Mode 'Compare Claude vs Mistral'\n3. Lancer la génération",
             "Rapport comparatif affiché : latence, coût, richesse, gagnant, recommandation. Toggle vue Claude/Mistral"),
            ("P2","[NEW] Multi-pays Sourcing IA (CI + SN + GH)",
             "Cocher plusieurs pays cibles dans le sélecteur",
             "","1. Sélecteur pays : cocher CI, SN, GH\n2. Plateformes pré-suggérées s'adaptent\n3. Générer",
             "L'IA produit des profils des 3 pays avec noms, localisations et plateformes locales cohérentes"),
            ("P2","[NEW] Contact email depuis carte profil",
             "Bouton Message ouvre un dialog avec subject + body pré-remplis",
             "","1. Carte profil → bouton 'Message'\n2. Vérifier le contenu pré-rempli\n3. Copier le message ou ouvrir LinkedIn search",
             "Dialog affiche objet/corps avec prénom/poste/entreprise. Lien LinkedIn search ouvre Google search préformatée"),
            ("P1","[FIX] Modal Nouvelle offre — pas de débordement",
             "Régression UI résolue : modal en flex-column avec body scrollable + footer sticky",
             "","1. Cliquer 'Nouvelle offre'\n2. Redimensionner navigateur à 768px de haut\n3. Vérifier que les boutons Annuler/Créer restent visibles",
             "Modal centrée, body scroll interne, header + footer toujours visibles, pas de débordement vertical"),
        ]
    },
    "SOURCING_IA": {
        "label": "Sourcing IA Multi-Pays Afrique",
        "color": "059669",
        "cases": [
            ("P1","Endpoint POST /jobs/:id/source — Claude seul",
             "Génération de profils synthétiques via Claude",
             "ANTHROPIC_API_KEY configuré, offre ouverte",
             "1. POST /api/recruitment/jobs/{id}/source\n2. Body: { model:'claude', countries:['CI'], platforms:['LinkedIn','Africawork'], max_profiles:5 }",
             "200 OK, data.profiles[] de 5 profils + strategy. meta.provider='claude', estimatedCostEur calculé. Audit log inséré"),
            ("P1","Endpoint POST /jobs/:id/source/compare — Claude vs Mistral",
             "Appels parallèles + métriques",
             "Les 2 clés IA configurées",
             "1. POST /api/recruitment/jobs/{id}/source/compare\n2. Body: { countries:['CI','SN'], max_profiles:3 }",
             "200 OK, comparison.winner = claude|mistral. ratios.latency/cost/richness retournés. recommendation textuelle"),
            ("P1","Rate-limit sourcing — 6 requêtes/min",
             "Protection coût IA (OWASP A05)",
             "","1. Boucler 7 fois POST /jobs/:id/source en moins d'1 minute",
             "Les 6 premières passent, la 7e retourne 429 Too Many Requests"),
            ("P1","Endpoint GET /jobs/:id/sourced-profiles",
             "Liste cache des profils générés pour une offre",
             "Profils en cache (seed ou générés)",
             "1. GET /api/recruitment/jobs/{id}/sourced-profiles",
             "200 OK, tableau ordonné par match_score DESC. Profils transférés et non-transférés inclus, distinguables par transferred_to_application_id"),
            ("P1","Endpoint transfer-all transactionnel",
             "Tous les profils non transférés transférés en une transaction",
             "≥1 profil non transféré","1. POST /api/recruitment/jobs/{id}/sourced-profiles/transfer-all\n2. Vérifier en DB",
             "Réponse data.transferred = N. Chaque profil a transferred_to_application_id. N candidatures créées (source=sourced_ai)"),
            ("P2","Idempotence transfert unique (409)",
             "Tenter de transférer 2x le même profil",
             "Profil déjà transféré","1. POST /api/recruitment/jobs/{id}/sourced-profiles/{profileId}/transfer\n2. Re-POST sur le même profileId",
             "1er appel: 201 Created avec applicationId. 2e appel: 409 Conflict avec applicationId existant"),
            ("P2","RBAC sourcing — employee 403",
             "Un employee ne peut pas lancer un sourcing",
             "Token employe@sotra.ci","1. POST /jobs/:id/source avec token employee","403 Forbidden"),
            ("P1","Devise auto par pays cible",
             "Le prompt et les profils utilisent la bonne devise selon le pays",
             "","1. Lancer sourcing avec countries=['NG']\n2. Vérifier les estimatedSalary",
             "Profils retournés avec estimatedSalaryCurrency='NGN' (Nigeria). Pour countries=['CI'] → XOF. Pour CM/TD → XAF"),
            ("P1","Score de richesse (computeSourcingRichness)",
             "Métrique 0-100 sur la qualité de la réponse IA",
             "","1. Générer un sourcing\n2. Vérifier meta.richnessScore","Score >0 si profils + strategy. ≥70 si réponse riche (5+ profils, bestPlatforms 2+, boolean search, salary benchmark)"),
        ]
    },
    "MULTI_PAYS_FILIALES": {
        "label": "Multi-pays & Filiales",
        "color": "7C3AED",
        "cases": [
            ("P1","Activation tous packs côté super_admin",
             "Onglet Multi-législatif affiche les 11 pays actifs",
             "Super admin connecté","1. /platform/settings\n2. Onglet 'Multi-législatif'\n3. Compter les cards 'Actif'",
             "11 cartes vert/Actif : CI/SN/BJ/TG/BF/ML/NE (UEMOA), CM/TD (CEMAC), NG/GH (CEDEAO). Drapeaux + badges zone + compteurs"),
            ("P1","Toggle has_subsidiaries sur tenant",
             "Activer l'option filiales sur un tenant existant",
             "Super admin","1. /platform/tenants/{id}\n2. Cocher 'A des filiales'\n3. Sauver",
             "platform.tenants.has_subsidiaries = true. Le tenant peut désormais créer plusieurs legal_entities avec pays différents"),
            ("P1","Création filiale multi-pays",
             "Admin tenant crée une filiale avec pays + pack législatif",
             "Admin SOTRA connecté","1. Paramètres → Entités juridiques\n2. Nouvelle entité 'SOTRA Dakar'\n3. Pays: Sénégal, pack auto sn_2024\n4. CNPS, RCCM, AT 2%\n5. Créer",
             "Filiale créée, country_code='SEN', legislation_pack_code='sn_2024'. Badge 'Filiale' visible si plusieurs entités"),
            ("P2","Édition filiale existante",
             "Modifier les infos d'une filiale (bouton ⚙)",
             "≥1 filiale créée","1. Card filiale → bouton ⚙\n2. Changer le pack législatif et l'AT\n3. Enregistrer",
             "PATCH /settings/legal-entities/{id} 200 OK. Mise à jour visible dans la card"),
            ("P1","Rattachement employé → filiale",
             "Champ legal_entity_id dans la fiche employé",
             "≥2 filiales, 1 employé","1. /employees/{id}\n2. Champ 'Filiale' → sélectionner 'SOTRA Bouaké'\n3. Sauver",
             "employees.legal_entity_id mis à jour. Compteur employés sur la card filiale incrémenté"),
            ("P1","Suppression filiale avec employés actifs → 409",
             "Garde-fou : impossible de supprimer une filiale avec employés",
             "Filiale avec ≥1 employé actif","1. Tenter DELETE /settings/legal-entities/{id}",
             "409 Conflict avec message 'Cette entite a des employes actifs'"),
            ("P2","Bloc d'aide explique le rattachement",
             "Pédagogie sur le mécanisme",
             "Onglet Entités juridiques","1. /settings → Entités juridiques\n2. Lire le bloc bleu en haut",
             "Texte explicite : les employés sont rattachés via legal_entity_id, le moteur de paie applique le pack législatif de la filiale"),
            ("P2","Devise locale par filiale",
             "L'UI affiche la devise du pack (XOF / XAF / NGN / GHS)",
             "Filiale créée hors zone XOF","1. Créer filiale Cameroun → vérifier card affiche XAF\n2. Filiale Nigeria → NGN",
             "La carte filiale affiche '(XAF)' ou '(NGN)' à côté du pack législatif"),
            ("P2","Compteur employés par filiale",
             "Card filiale affiche le nombre d'employés rattachés",
             "","1. Rattacher 5 employés à une filiale\n2. Recharger Settings","Icône Users + 5 sur la card"),
            ("P3","Pré-remplissage pack législatif depuis pays",
             "Quand on change le pays dans le formulaire, le pack se met à jour automatiquement",
             "","1. Modal Nouvelle filiale\n2. Changer pays de CIV vers SEN","Champ Pack législatif passe de 'ci_2024' à 'sn_2024' automatiquement"),
            ("P3","Wizard tenant — sélecteur pays par défaut",
             "Création de tenant propose les pays disponibles",
             "Super admin","1. /platform/tenants/new\n2. Étape 1 → champ Pays\n3. Vérifier les options",
             "Liste affiche au moins CIV/SEN/BFA/MLI/TGO/BEN/NER/GNB"),
        ]
    },
    "FORMATION": {
        "label": "Formation & FDFP",
        "color": "1D4ED8",
        "cases": [
            ("P1","Catalogue formations — affichage","",
             "5+ formations seedées","1. /training\n2. Observer les cards",
             "Cards avec titre, durée (heures), format (présentiel/e-learning), places disponibles"),
            ("P2","Création formation (admin)","",
             "Admin connecté","1. /training/new\n2. Remplir titre, description, durée, format, places max\n3. Cocher FDFP éligible si applicable",
             "Formation créée, visible dans le catalogue"),
            ("P1","Session planifiée — dates futures","",
             "Formation créée","1. Créer une session avec dates futures\n2. Définir lieu et formateur",
             "Session visible avec places disponibles décomptées dynamiquement"),
            ("P1","Inscription self-service — employee","",
             "employe@sotra.ci connecté, places disponibles","1. /mon-espace/formation\n2. Catalogue → S'inscrire",
             "Inscription créée, places disponibles -1, confirmation affichée"),
            ("P2","Vérification places avant inscription","",
             "Session complète (0 place)","1. Tenter de s'inscrire à une session pleine",
             "Erreur 'Session complète', bouton désactivé"),
            ("P2","Désinscription","",
             "Inscrit à une session future","1. Mes inscriptions → Annuler",
             "Inscription supprimée, places disponibles +1"),
            ("P3","Attestation téléchargeable (formation terminée)","",
             "Formation passée, statut completed","1. Mes inscriptions → Télécharger attestation",
             "PDF attestation téléchargé avec nom, date, formation, durée"),
            ("P3","Filtre formations agréées FDFP","",
             "","1. Catalogue → Filtre 'FDFP éligible'",
             "Seules les formations avec is_fdfp_eligible=true affichées"),
            ("P3","Demande remboursement FDFP","",
             "Formation agréée terminée","1. POST /training/fdfp/request\n2. Dossier complet",
             "Demande créée, statut 'pending FDFP'"),
        ]
    },
    "FRAIS": {
        "label": "Notes de Frais",
        "color": "BE185D",
        "cases": [
            ("P1","Création note de frais — brouillon","",
             "Employé connecté","1. /mon-espace/notes-de-frais → Nouvelle note\n2. Titre, mois\n3. Sauvegarder brouillon",
             "Note créée statut 'draft', visible dans la liste"),
            ("P1","Ajout lignes à la note","Description, catégorie, montant FCFA, justificatif",
             "Note brouillon","1. Ouvrir la note\n2. + Ajouter ligne\n3. Repas 8 500 FCFA\n4. Upload reçu",
             "Ligne ajoutée, total mis à jour automatiquement en FCFA"),
            ("P1","Soumission note → statut 'submitted'","",
             "Note avec au moins 1 ligne","1. Note brouillon → Soumettre",
             "Statut = submitted, manager notifié, modification impossible"),
            ("P1","Validation manager → statut 'approved'","",
             "Note submitted + manager connecté","1. Manager → Notes à valider → Approuver",
             "Statut = approved, RH peut procéder au remboursement"),
            ("P2","Refus manager → statut 'rejected'","",
             "","1. Manager → Refuser avec motif",
             "Statut = rejected, employé notifié avec motif"),
            ("P2","Remboursement via Mobile Money","",
             "Note approuvée, numéro Mobile Money enregistré","1. RH → Rembourser → Wave\n2. Confirmer montant",
             "Paiement Mobile Money initié, statut = remboursé, référence TXN stockée"),
            ("P2","Sauvegarde brouillon autosave","",
             "","1. Modifier une note\n2. Ne pas cliquer Sauvegarder\n3. Quitter et revenir",
             "Modifications conservées (autosave ou confirmation de quitter)"),
            ("P1","Employee voit uniquement ses propres notes","",
             "","1. employe@sotra.ci → /mon-espace/notes-de-frais",
             "Seules ses propres notes affichées, jamais celles des collègues"),
        ]
    },
    "CARRIERES": {
        "label": "Carrières & Compétences",
        "color": "6D28D9",
        "cases": [
            ("P1","Référentiel compétences — création","",
             "Admin/RH connecté","1. /careers/skills → Nouvelle compétence\n2. Nom, catégorie (technique/comportemental/managérial)",
             "Compétence créée dans le référentiel"),
            ("P2","Attribution compétences à un employé","",
             "","1. Fiche employé → Compétences\n2. Sélectionner compétence + niveau (1-5)\n3. Sauvegarder",
             "Compétence liée à l'employé avec niveau"),
            ("P1","Entretien annuel — création et saisie","",
             "","1. /careers/evaluations → Nouvel entretien\n2. Sélectionner employé, année\n3. Saisir scores performance et compétences",
             "Entretien créé statut 'completed' avec scores et commentaires"),
            ("P2","Historique entretiens par employé","",
             "","1. Fiche employé → Entretiens",
             "Liste des entretiens triée par année, scores visibles"),
            ("P3","Vue 9-Box","Performance vs Potentiel",
             "Évaluations avec scores","1. /careers/9box",
             "Grille 9-box affichée avec positionnement des employés selon scores"),
            ("P3","Score rétention IA","Calcul du risque de départ",
             "ANTHROPIC_API_KEY configuré","1. /careers/retention → Analyser",
             "Score + niveau de risque (low/medium/high) + facteurs + recommandations en français"),
        ]
    },
    "CNPS": {
        "label": "Déclarations CNPS & DISA",
        "color": "065F46",
        "cases": [
            ("P1","Génération déclaration mensuelle CNPS","",
             "Bulletins clôturés du mois","1. /cnps/declarations → Générer pour Mai 2025",
             "Déclaration créée avec total cotisations salariales et patronales en FCFA"),
            ("P1","Export e-CNPS CSV format officiel","",
             "Déclaration générée","1. Déclaration → Exporter CSV",
             "CSV téléchargé compatible plateforme e-CNPS CI, format correct"),
            ("P2","Alerte date limite déclaration (avant le 15)","",
             "En date du 10 du mois","1. Observer le dashboard admin",
             "Alerte 'Déclaration CNPS à soumettre avant le 15' visible"),
            ("P1","Génération DISA annuelle","Déclaration Individuelle Salaires Annuels",
             "12 mois de bulletins","1. /cnps/disa → Générer DISA 2025",
             "DISA créée pour chaque employé avec cumuls annuels salaire brut + cotisations + ITS"),
            ("P2","Export DISA (CSV/XML)","",
             "DISA générée","1. DISA → Exporter",
             "Fichier DISA téléchargé au format requis par CNPS/DGI"),
            ("P1","RNS PDF généré — formulaire EN-GDAV-06 v03","Relevé Nominatif des Salaires",
             "rns-template.pdf copié dans dist/assets (fix Dockerfile)","1. /cnps/rns → GET /cnps/rns/2026/pdf",
             "PDF téléchargé avec données employeur, liste employés, salaires annuels en FCFA"),
            ("P2","RNS filtré par employé","",
             "","1. GET /cnps/rns/2026/pdf?employeeId=XXX",
             "PDF généré pour un seul employé"),
            ("P3","Attestation conformité CNPS","",
             "","1. /cnps/employers/certificate",
             "Attestation PDF téléchargée"),
        ]
    },
    "MOBILE_MONEY": {
        "label": "Paiements Mobile Money",
        "color": "B45309",
        "cases": [
            ("P1","Création campagne virement salaires","",
             "Bulletins générés + clôturés","1. /payroll/mobile-money/campaigns → Nouvelle campagne\n2. Sélectionner période",
             "Campagne créée avec liste des 80 virements à effectuer"),
            ("P2","Exécution campagne — Wave","",
             "Campagne créée, API Wave configurée","1. Campagne → Exécuter\n2. Confirmer",
             "Virements Wave initiés, statuts = pending"),
            ("P2","Exécution campagne — MTN MoMo","",
             "","1. Filtrer employés MTN\n2. Exécuter",
             "Virements MTN initiés"),
            ("P2","Exécution campagne — Orange Money","",
             "","1. Filtrer employés Orange\n2. Exécuter",
             "Virements Orange initiés"),
            ("P2","Vérification numéro avant virement","POST /payroll/mobile-money/verify-number",
             "","1. Saisir un numéro +225 07 XX XX XX\n2. Vérifier",
             "Numéro actif confirmé ou erreur explicite"),
            ("P2","Callback webhook → mise à jour statut","",
             "Virement initié","1. Simuler callback success du provider\n2. Vérifier statut",
             "Statut mis à jour de 'pending' à 'success'"),
            ("P2","Virement failed → alerte","",
             "","1. Simuler callback failed",
             "Statut = failed, alerte admin, virement visible dans liste des échecs"),
            ("P1","Bulletin marqué 'payé' après virement success","",
             "","1. Virement success\n2. Ouvrir le bulletin correspondant",
             "payment_status = 'paid', référence TXN affichée sur le bulletin"),
            ("P1","Validation format téléphone CI","",
             "","1. Saisir +225 07 12 34 56 78 (Wave valide)\n2. Saisir +33 6 XX (invalide)",
             "Format +225 07/05 + 8 chiffres accepté, autres rejetés avec erreur claire"),
        ]
    },
    "IA": {
        "label": "Assistant IA (Claude CI)",
        "color": "7C3AED",
        "cases": [
            ("P1","Chat IA — réponse en streaming SSE","",
             "ANTHROPIC_API_KEY configuré","1. /ai/chat\n2. Poser une question RH\n3. Observer la réponse",
             "Réponse arrivant en streaming (SSE), affichée progressivement"),
            ("P1","Réponse toujours en français","",
             "","1. Poser la question en français\n2. (Optionnel) tenter en anglais",
             "Réponse 100% en français, références légales CI citées"),
            ("P2","Question CNPS → calcul correct","",
             "","1. 'Quel est le taux CNPS retraite salarié ?'",
             "Réponse: 6.3% avec mention du plafond 1 647 315 FCFA"),
            ("P2","Question congés CI","",
             "","1. 'Combien de jours de congés pour 8 mois ?'",
             "Réponse: 8 × 2.5 = 20 jours ouvrables (Code du Travail CI)"),
            ("P2","Question taux AT par secteur","",
             "","1. 'Quel est le taux AT pour une entreprise BTP ?'",
             "Réponse: 3% (BTP/Transport) avec source réglementaire"),
            ("P3","Génération CDI CI — OHADA","",
             "","1. IA → Générer contrat → CDI CI\n2. Saisir infos employé",
             "PDF CDI CI avec clauses OHADA obligatoires, NNI, CNPS, période d'essai légale CI"),
            ("P3","Génération certificat de travail","",
             "","1. IA → Générer document → Certificat de travail",
             "Document complet avec données employé, poste, dates, mentions légales CI"),
            ("P3","Analyse rétention — score et facteurs","",
             "","1. /ai/retention → Analyser pour un employé",
             "{ score, risk: low|medium|high, factors[], recommendations[] } en français"),
            ("P2","Employee — accès limité à l'IA","",
             "employe@sotra.ci connecté","1. Accéder au chat IA\n2. Poser des questions avancées",
             "Questions simples autorisées, génération de documents refusée (rôle limité)"),
            ("P2","Contexte tenant injecté dans le prompt","",
             "","1. Demander 'Quel est mon secteur d'activité ?'",
             "IA répond avec le secteur du tenant connecté (ex: 'Transport pour SOTRA')"),
        ]
    },
    "REPORTING": {
        "label": "Reporting & Analytics",
        "color": "0E7490",
        "cases": [
            ("P1","KPIs dashboard RH — valeurs correctes","Effectifs, masse salariale FCFA, cotisations CNPS, postes ouverts",
             "Données seedées","1. Connecter admin\n2. Observer /dashboard",
             "Valeurs cohérentes avec la base de données (ex: 80 employés SOTRA)"),
            ("P1","Graphique évolution effectifs 12 mois","LineChart",
             "","1. Observer graphique dashboard",
             "Courbe affichée, points par mois, valeurs cohérentes"),
            ("P2","Répartition départements — BarChart","",
             "","1. Observer BarChart dashboard",
             "Barres par département avec nombre d'employés, ordonné"),
            ("P2","Cotisations CNPS mensuelles","Total patronal + salarial du mois",
             "","1. KPI card CNPS\n2. Vérifier la valeur du mois en cours",
             "Valeur = somme des cotisations des bulletins du mois sélectionné"),
            ("P2","Taux absentéisme calculé","% jours d'absence / jours ouvrables",
             "","1. KPI card absentéisme",
             "Taux affiché en %, calculé sur les absences approuvées"),
            ("P2","Export rapport masse salariale","",
             "","1. Reporting → Masse salariale → Exporter\n2. Choisir période",
             "Excel/CSV téléchargé avec employé, brut, net, cotisations FCFA"),
            ("P2","Insight IA — alerte rétention","3 alertes max sur le dashboard",
             "IA configurée","1. Observer la section 'Insights IA' du dashboard",
             "Maximum 3 alertes avec risque de départ, absentéisme, essais expirants"),
            ("P3","Alerte CNPS à déclarer sur dashboard","",
             "Après le 10 du mois","1. Observer le panneau alertes",
             "Alerte 'Déclaration CNPS à soumettre avant le 15' visible"),
        ]
    },
    "SETTINGS": {
        "label": "Paramètres Tenant",
        "color": "374151",
        "cases": [
            ("P2","Modifier logo tenant","Upload logo pour personnalisation",
             "Admin connecté","1. Paramètres → Apparence → Upload logo\n2. Choisir image PNG\n3. Sauvegarder",
             "Logo affiché dans la sidebar et sur la page login"),
            ("P2","Modifier couleur primaire/secondaire","",
             "","1. Paramètres → Couleur primaire → Changer\n2. Sauvegarder\n3. Rafraîchir",
             "Nouvelle couleur appliquée à toute l'interface"),
            ("P1","Gestion utilisateurs tenant — liste","Admin peut voir tous les users du tenant",
             "","1. Paramètres → Utilisateurs",
             "Liste avec email, rôle, statut, date création"),
            ("P2","Invitation nouvel utilisateur","",
             "","1. + Inviter\n2. Email, rôle (dropdown)\n3. Envoyer",
             "Email d'invitation envoyé, token d'invitation créé"),
            ("P2","Modification rôle utilisateur","",
             "","1. Utilisateur → Modifier rôle → hr_officer",
             "Rôle mis à jour, permissions ajustées immédiatement"),
            ("P2","Désactivation utilisateur tenant","",
             "","1. Utilisateur → Désactiver",
             "is_active = false, cet user ne peut plus se connecter"),
        ]
    },
    "RBAC": {
        "label": "Contrôle d'accès (RBAC)",
        "color": "7F1D1D",
        "cases": [
            ("P1","super_admin ne peut pas accéder /dashboard","Redirection /platform/dashboard",
             "superadmin connecté","1. Naviguer vers /dashboard",
             "Redirection automatique vers /platform/dashboard, pas d'erreur"),
            ("P1","admin d'un tenant ne peut pas accéder /platform/*","403 ou redirection",
             "admin@sotra.ci connecté","1. Tenter /platform/tenants",
             "403 Forbidden ou redirection /dashboard, jamais l'interface super admin"),
            ("P1","employee tente /employees → redirigé /mon-espace","Guard RoleGuard actif",
             "employe@sotra.ci","1. URL directe /employees",
             "Redirection automatique /mon-espace"),
            ("P1","hr_officer ne peut pas clôturer la paie","Action non disponible dans l'UI",
             "hr_officer connecté","1. Accéder /payroll\n2. Chercher le bouton Clôturer",
             "Bouton absent ou désactivé, API retourne 403 si tenté directement"),
            ("P1","manager ne voit que son équipe","Filtrage automatique",
             "manager@sotra.ci lié à Exploitation","1. Accéder /employees depuis manager",
             "Seuls les 30 employés Exploitation listés (pas les 80 SOTRA)"),
            ("P2","readonly — aucune modification possible","Tous les boutons d'édition absents",
             "Utilisateur rôle readonly","1. Connecter\n2. Tenter toute modification",
             "403 sur toutes les routes PUT/POST/DELETE, UI en lecture seule"),
            ("P1","Isolation multi-tenant — token cross-tenant","Token SOTRA ne donne pas accès aux données Cabinet",
             "Token SOTRA JWT","1. Utiliser le token SOTRA pour appeler /api/employees avec schéma cabinet",
             "403 ou données du bon schéma uniquement, jamais de fuite cross-tenant"),
        ]
    },
    "MON_ESPACE": {
        "label": "Espace Employee (Self-Service)",
        "color": "065F46",
        "cases": [
            ("P1","Dashboard /mon-espace — soldes congés affichés","Barres de progression colorées",
             "employe@sotra.ci connecté, soldes seedés","1. Connecter employe\n2. Observer /mon-espace",
             "CP, Maladie affichés avec jours acquis/pris/restants, barres colorées"),
            ("P1","Badge 'Nouveau' bulletin — /mon-espace","",
             "Bulletin non consulté","1. Observer /mon-espace après génération d'un nouveau bulletin",
             "Badge 'Nouveau' visible sur le bulletin le plus récent"),
            ("P1","Sidebar réduite — 5 items seulement","Mon espace, Absences, Bulletins, Frais, Formation, Profil",
             "employe connecté","1. Observer la sidebar",
             "Exactement 6 items, aucun lien vers /employees, /payroll global, /recruitment"),
            ("P2","Mon profil — modification téléphone","",
             "","1. /mon-espace/profil → Modifier téléphone\n2. Sauvegarder",
             "Téléphone mis à jour, Mobile Money provider visible"),
            ("P2","Mon profil — changement de mot de passe","",
             "","1. Saisir ancien mot de passe\n2. Nouveau + confirmation\n3. Sauvegarder",
             "Mot de passe changé, session maintenue"),
            ("P2","Mes 3 dernières absences affichées","Sur le dashboard /mon-espace",
             "","1. Observer la section absences du dashboard",
             "3 dernières absences avec badge statut coloré"),
            ("P2","Formations recommandées — 2 cards","",
             "","1. Observer la section formation du dashboard",
             "2 cards de formations affichées"),
            ("P1","Accès /mon-espace/bulletins","Liste des 24 derniers bulletins",
             "","1. /mon-espace/bulletins",
             "Liste avec mois, net payable FCFA, statut, bouton télécharger PDF"),
        ]
    },
    "E2E": {
        "label": "Scénarios End-to-End",
        "color": "1E3A5F",
        "cases": [
            ("P1","Flux admin complet — de zéro à bulletins clôturés",
             "Créer employé → créer période → générer bulletins → clôturer → télécharger PDF",
             "Admin SOTRA connecté",
             "1. Créer nouvel employé avec salaire 250 000 FCFA\n2. /payroll → Nouvelle période Mai 2025\n3. Générer bulletins\n4. Vérifier calcul CNPS+ITS\n5. Clôturer\n6. Télécharger PDF",
             "Tout le flux sans erreur 404/500, PDF correct avec données CI"),
            ("P1","Flux employé self-service complet",
             "Connexion → demande absence → consultation bulletin → note de frais",
             "employe@sotra.ci",
             "1. Connexion → /mon-espace\n2. Nouvelle demande absence CP 3 jours\n3. Consulter dernier bulletin → Télécharger\n4. Créer note de frais brouillon",
             "Toutes les actions réussies sans erreur, données sauvegardées"),
            ("P1","Flux RH — recrutement jusqu'à contrat",
             "Offre → candidature → hired → contrat",
             "hr_manager connecté",
             "1. Créer offre 'Technicien'\n2. Ajouter candidature\n3. Avancer jusqu'à 'hired'\n4. Générer contrat CDI CI",
             "Contrat PDF généré avec clauses OHADA"),
            ("P1","Flux paie CNPS complet",
             "Bulletins → déclaration → export e-CNPS",
             "Admin + bulletins clôturés",
             "1. Clôturer paie d'un mois\n2. /cnps → Générer déclaration\n3. Exporter CSV e-CNPS\n4. Vérifier totaux cohérents avec bulletins",
             "CSV cohérent avec la somme des cotisations des bulletins"),
            ("P1","Création tenant via portail → login admin → données",
             "Super admin crée tenant, admin se connecte et voit les données démo",
             "Super admin connecté",
             "1. Créer tenant 'TestCo' avec seedDemoData=true\n2. Noter tempPassword\n3. Se connecter avec admin créé\n4. Observer /dashboard",
             "Dashboard RH avec 8 employés, bulletins, absences — aucun écran vide"),
            ("P2","Flux paiement Mobile Money — campagne Wave complète",
             "Clôture → campagne → exécution → webhook → bulletin payé",
             "Admin + bulletins clôturés + Wave configuré",
             "1. Clôturer paie\n2. Créer campagne Mobile Money\n3. Exécuter pour employés Wave\n4. Simuler callback success\n5. Vérifier bulletin marqué 'payé'",
             "TXN référence stockée, bulletin payment_status = 'paid'"),
        ]
    },
}

COLUMN_HEADERS = [
    ("ID Test", 12),
    ("Priorité", 13),
    ("Titre du scénario", 38),
    ("Description / Contexte", 32),
    ("Prérequis / Données", 30),
    ("Étapes détaillées", 40),
    ("Résultat attendu", 38),
    ("Résultat obtenu", 38),
    ("Statut", 18),
    ("Testeur", 16),
    ("Date test", 13),
    ("Durée (min)", 12),
    ("Ticket / Requalification", 24),
    ("Commentaire / Amélioration", 36),
    ("Sprint", 10),
    ("Env.", 12),
]

MODULE_ABBR = {
    "AUTH": "AUTH", "PLATFORM": "PLT", "EMPLOYEES": "EMP",
    "PAYROLL": "PAY", "ABSENCES": "ABS", "RECRUTEMENT": "REC",
    "SOURCING_IA": "SRC", "MULTI_PAYS_FILIALES": "MPF",
    "FORMATION": "FRM", "FRAIS": "FRA", "CARRIERES": "CAR",
    "CNPS": "CNP", "MOBILE_MONEY": "MM", "IA": "AI",
    "REPORTING": "REP", "SETTINGS": "SET", "RBAC": "RBA",
    "MON_ESPACE": "ESP", "E2E": "E2E",
}

PRIO_COLORS = {
    "P1 — Critique": ("FEE2E2", "DC2626"),
    "P2 — Haute":    ("FEF9C3", "D97706"),
    "P3 — Moyenne":  ("DBEAFE", "2563EB"),
    "P4 — Basse":    ("F0FDF4", "16A34A"),
}

STATUS_COLORS = {
    "✅ Passé":        ("DCFCE7", "166534"),
    "❌ Échoué":       ("FEE2E2", "7F1D1D"),
    "⚠️ Bloqué":       ("FEF9C3", "78350F"),
    "▶️ En cours":     ("DBEAFE", "1E40AF"),
    "⬜ Non exécuté":  ("F8FAFC", "64748B"),
}

def make_module_sheet(wb, module_key, module_data):
    abbr   = MODULE_ABBR[module_key]
    color  = module_data["color"]
    label  = module_data["label"]
    cases  = module_data["cases"]
    sheet_name = abbr

    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Row 1 — Module header
    ws.row_dimensions[1].height = 26
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMN_HEADERS))}1")
    c = ws.cell(row=1, column=1,
                value=f"  {label.upper()}  ·  {len(cases)} scénarios")
    c.fill   = PatternFill("solid", fgColor=color)
    c.font   = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2 — Column headers
    ws.row_dimensions[2].height = 22
    for ci, (hdr, width) in enumerate(COLUMN_HEADERS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        c = ws.cell(row=2, column=ci, value=hdr)
        c.fill      = PatternFill("solid", fgColor="1E3A5F")
        c.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

    # Rows 3+ — Test cases
    for ri, (prio, title, desc, prereq, steps, expected) in enumerate(cases, 3):
        row_num = ri
        ws.row_dimensions[row_num].height = 58

        prio_full = {"P1": "P1 — Critique", "P2": "P2 — Haute",
                     "P3": "P3 — Moyenne", "P4": "P4 — Basse"}[prio]
        test_id   = f"{abbr}-{(ri - 2):03d}"
        bg        = C_BGLIGHT if ri % 2 == 0 else C_WHITE

        vals = [test_id, prio_full, title, desc, prereq, steps, expected,
                "", "⬜ Non exécuté", "", "", "", "", "", "Sprint 1", "Production"]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border    = thin_border()
            c.font      = Font(size=9, name="Calibri")

            # ID col
            if ci == 1:
                c.fill = PatternFill("solid", fgColor=color)
                c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            # Priority col
            elif ci == 2:
                pb, pf = PRIO_COLORS.get(prio_full, ("FFFFFF", "000000"))
                c.fill = PatternFill("solid", fgColor=pb)
                c.font = Font(bold=True, color=pf, size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            # Status col
            elif ci == 9:
                c.fill = PatternFill("solid", fgColor="F8FAFC")
                c.font = Font(color="64748B", size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = PatternFill("solid", fgColor=bg)

        # Data validation — Status
        dv_status = DataValidation(type="list", formula1=STATUS_LIST, allow_blank=True)
        dv_status.sqref = f"I{row_num}"
        ws.add_data_validation(dv_status)

        # Data validation — Priority
        dv_prio = DataValidation(type="list", formula1=PRIO_LIST, allow_blank=True)
        dv_prio.sqref = f"B{row_num}"
        ws.add_data_validation(dv_prio)

        # Data validation — Env
        dv_env = DataValidation(type="list", formula1=ENV_LIST, allow_blank=True)
        dv_env.sqref = f"P{row_num}"
        ws.add_data_validation(dv_env)

    # Conditional formatting on Status column
    last_row = 2 + len(cases)
    for status, (bg_c, fg_c) in STATUS_COLORS.items():
        rule = FormulaRule(
            formula=[f'$I3="{status}"'],
            fill=PatternFill("solid", fgColor=bg_c),
            font=Font(color=fg_c, size=9, name="Calibri")
        )
        ws.conditional_formatting.add(f"A3:P{last_row}", rule)

    return ws


def make_dashboard(wb, logo_path):
    ws = wb.active
    ws.title = "DASHBOARD"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    # ── Logo ──
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width  = 160
        img.height = 60
        ws.add_image(img, "A1")
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 20

    # ── Title block ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14

    ws.merge_cells("C1:I2")
    c = ws.cell(row=1, column=3,
                value="NexusRH CI — Cahier de Recettes Officiel\nOpenLab Consulting · v1.0")
    c.font      = Font(bold=True, size=18, color="1E3A5F", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("C3:I3")
    c = ws.cell(row=3, column=3, value=f"Généré le {TODAY}  ·  Couverture cible : 100% des fonctionnalités")
    c.font      = Font(size=10, color="64748B", name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Section header helper ──
    def section_header(row, title):
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f"A{row}:I{row}")
        c = ws.cell(row=row, column=1, value=f"  {title}")
        c.fill      = PatternFill("solid", fgColor="1E3A5F")
        c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")

    # ── Summary table ──
    section_header(5, "RÉSUMÉ PAR MODULE")
    hdrs = ["Module", "Total", "✅ Passé", "❌ Échoué", "⚠️ Bloqué", "▶️ En cours", "⬜ Non exécuté", "% Réussite", "Statut"]
    hdr_colors = [C_NAVY, "374151", "166534", "7F1D1D", "78350F", "1E40AF", "374151", "374151", "374151"]
    ws.row_dimensions[6].height = 20

    for ci, (hdr, hc) in enumerate(zip(hdrs, hdr_colors), 1):
        c = ws.cell(row=6, column=ci, value=hdr)
        c.fill      = PatternFill("solid", fgColor=hc)
        c.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    MODULE_ORDER = [
        ("AUTH", "Authentification"),
        ("PLT", "Portail Super Admin"),
        ("EMP", "Employés"),
        ("PAY", "Paie CI (CNPS+ITS)"),
        ("ABS", "Absences"),
        ("REC", "Recrutement"),
        ("FRM", "Formation / FDFP"),
        ("FRA", "Notes de Frais"),
        ("CAR", "Carrières"),
        ("CNP", "CNPS & DISA"),
        ("MM",  "Mobile Money"),
        ("AI",  "Assistant IA"),
        ("REP", "Reporting"),
        ("SET", "Paramètres"),
        ("RBA", "RBAC"),
        ("ESP", "Espace Employé"),
        ("E2E", "End-to-End"),
    ]

    data_rows = {}
    row = 7
    for abbr, label in MODULE_ORDER:
        ws.row_dimensions[row].height = 18
        data_rows[abbr] = row

        # Find the sheet
        sheet_names = [s.title for s in wb.worksheets]
        if abbr in sheet_names:
            sn = abbr
        else:
            sn = abbr

        # Module name
        c = ws.cell(row=row, column=1, value=label)
        c.font      = Font(bold=True, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = thin_border()
        c.fill      = PatternFill("solid", fgColor=C_BGLIGHT if row % 2 == 0 else C_WHITE)

        # Total (COUNTA on ID column, minus header)
        ws.cell(row=row, column=2,
                value=f"=COUNTA('{sn}'!A:A)-2").border = thin_border()
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2).font = Font(bold=True, size=9, name="Calibri")

        for col_offset, status in enumerate(["✅ Passé","❌ Échoué","⚠️ Bloqué","▶️ En cours","⬜ Non exécuté"], 3):
            c2 = ws.cell(row=row, column=col_offset,
                         value=f'=COUNTIF(\'{sn}\'!I:I,"{status}")')
            c2.border    = thin_border()
            c2.alignment = Alignment(horizontal="center")
            c2.font      = Font(size=9, name="Calibri")

        # % Réussite
        c3 = ws.cell(row=row, column=8,
                     value=f"=IF(B{row}=0,0,C{row}/B{row})")
        c3.number_format = "0%"
        c3.border    = thin_border()
        c3.alignment = Alignment(horizontal="center")
        c3.font      = Font(bold=True, size=9, name="Calibri")

        # Statut indicateur
        c4 = ws.cell(row=row, column=9,
                     value=f'=IF(H{row}=1,"🟢 Complet",IF(H{row}>0.7,"🟡 Avancé",IF(H{row}>0,"🟠 En cours","⬜ À faire")))')
        c4.border    = thin_border()
        c4.alignment = Alignment(horizontal="center")
        c4.font      = Font(size=9, name="Calibri")

        row += 1

    # Total row
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=1, value="TOTAL GLOBAL")
    c.fill   = PatternFill("solid", fgColor=C_NAVY)
    c.font   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c.border = thick_border()

    for col in range(2, 9):
        c2 = ws.cell(row=row, column=col,
                     value=f"=SUM({get_column_letter(col)}7:{get_column_letter(col)}{row-1})")
        if col == 8:
            c2.value = f"=IF(B{row}=0,0,C{row}/B{row})"
            c2.number_format = "0%"
        c2.fill      = PatternFill("solid", fgColor=C_NAVY)
        c2.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        c2.alignment = Alignment(horizontal="center")
        c2.border    = thick_border()

    total_row = row

    # ── KPI cards (text-based) ──
    kpi_row = total_row + 3
    section_header(kpi_row, "INDICATEURS CLÉS EN TEMPS RÉEL")

    kpi_row += 1
    ws.row_dimensions[kpi_row].height = 50
    kpis = [
        ("Total scénarios", f"=B{total_row}", C_NAVY),
        ("✅ Passés",        f"=C{total_row}", C_GREEN),
        ("❌ Échoués",       f"=D{total_row}", C_RED),
        ("⚠️ Bloqués",       f"=E{total_row}", C_AMBER),
        ("▶️ En cours",      f"=F{total_row}", C_BLUE),
        ("⬜ Non exécutés",  f"=G{total_row}", C_SLATE),
        ("% Réussite",       f"=H{total_row}", C_GREEN),
        ("Couverture",       f"=TEXT(1-G{total_row}/B{total_row},\"0%\")", C_ORANGE),
    ]
    for ci, (lbl, formula, color) in enumerate(kpis, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width, 14
        )
        top = ws.cell(row=kpi_row, column=ci, value=lbl)
        top.fill      = PatternFill("solid", fgColor=color)
        top.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        top.alignment = Alignment(horizontal="center", vertical="bottom")
        top.border    = thick_border()

        bot = ws.cell(row=kpi_row + 1, column=ci, value=formula)
        if "%" in lbl:
            bot.number_format = "0%"
        bot.fill      = PatternFill("solid", fgColor="F8FAFC")
        bot.font      = Font(bold=True, size=18, color=color, name="Calibri")
        bot.alignment = Alignment(horizontal="center", vertical="center")
        bot.border    = thick_border()

    ws.row_dimensions[kpi_row + 1].height = 40

    # ── Charts ──
    chart_row = kpi_row + 4

    # Bar Chart — tests par module
    bar = BarChart()
    bar.type     = "col"
    bar.grouping = "clustered"
    bar.title    = "Statut des tests par module"
    bar.y_axis.title = "Nb scénarios"
    bar.x_axis.title = "Module"
    bar.width    = 18
    bar.height   = 12
    bar.style    = 10

    first_data_row = 7
    last_data_row  = total_row - 1
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)

    for col_idx, (status, color_hex) in enumerate([
        ("Passé", "16A34A"), ("Échoué", "DC2626"),
        ("Bloqué", "D97706"), ("En cours", "2563EB")
    ], 3):
        data = Reference(ws, min_col=col_idx, min_row=6,
                         max_row=last_data_row)
        bar.add_data(data, titles_from_data=True)
        bar.series[-1].graphicalProperties.solidFill = color_hex
        bar.series[-1].graphicalProperties.line.solidFill = color_hex

    bar.set_categories(cats)
    bar.shape = 4
    ws.add_chart(bar, f"A{chart_row}")

    # Pie Chart — distribution globale
    pie = PieChart()
    pie.title  = "Distribution globale des statuts"
    pie.width  = 12
    pie.height = 12
    pie.style  = 10

    pie_data = Reference(ws, min_col=3, max_col=7,
                         min_row=total_row, max_row=total_row)
    pie_labels = Reference(ws, min_col=3, max_col=7,
                           min_row=6, max_row=6)
    pie.add_data(pie_data)
    pie.set_categories(pie_labels)

    slice_colors = ["16A34A", "DC2626", "D97706", "2563EB", "94A3B8"]
    for i, hex_c in enumerate(slice_colors):
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill = hex_c
        pie.series[0].dPt.append(dp)

    ws.add_chart(pie, f"E{chart_row}")

    return ws


def make_guide(wb):
    ws = wb.create_sheet(title="GUIDE", index=1)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30

    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="MODE D'EMPLOI — NexusRH CI Cahier de Recettes")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    rows = [
        (3, "STATUTS DE TEST", "", ""),
        (4, "Statut", "Signification", "Couleur"),
        (5, "✅ Passé", "Test exécuté et résultat conforme", "Vert"),
        (6, "❌ Échoué", "Résultat non conforme — bug identifié", "Rouge"),
        (7, "⚠️ Bloqué", "Test non exécutable (dépendance manquante)", "Jaune"),
        (8, "▶️ En cours", "Test en cours d'exécution", "Bleu"),
        (9, "⬜ Non exécuté", "Test planifié, pas encore démarré", "Gris"),
        (11, "PRIORITÉS", "", ""),
        (12, "P1 — Critique", "Fonctionnalité bloquante, doit passer avant release", "Rouge foncé"),
        (13, "P2 — Haute", "Fonctionnalité importante, impact élevé si KO", "Amber"),
        (14, "P3 — Moyenne", "Fonctionnalité utile, contournement possible", "Bleu"),
        (15, "P4 — Basse", "Amélioration ou cas limite", "Vert"),
        (17, "REQUALIFICATION EN TICKET", "", ""),
        (18, "Colonne 'Ticket / Requalification'", "Saisir le numéro de ticket (ex: NRH-123)", ""),
        (19, "Commentaire / Amélioration", "Détailler le comportement observé et suggestion", ""),
        (21, "WORKFLOW DE TEST", "", ""),
        (22, "1. Sélectionner le module (onglet)", "Chaque onglet = un module fonctionnel", ""),
        (23, "2. Exécuter chaque scénario", "Suivre les étapes détaillées", ""),
        (24, "3. Mettre à jour le Statut (colonne I)", "Dropdown avec les 5 statuts", ""),
        (25, "4. Saisir Résultat obtenu (colonne H)", "Comportement réel observé", ""),
        (26, "5. Créer un ticket si Échoué/Bloqué", "Saisir référence dans colonne M", ""),
        (27, "6. Le DASHBOARD se met à jour automatiquement", "Formules COUNTIF dynamiques", ""),
        (29, "COMPTES DE TEST", "", ""),
        (30, "superadmin@nexusrh-ci.com / SuperAdmin1234!", "Super Admin Plateforme → /platform/dashboard", ""),
        (31, "admin@sotra.ci / Admin1234!", "Admin SOTRA → /dashboard (thème orange)", ""),
        (32, "rh@sotra.ci / Admin1234!", "RH Manager SOTRA → /dashboard", ""),
        (33, "manager@sotra.ci / Admin1234!", "Manager SOTRA → /dashboard équipe", ""),
        (34, "employe@sotra.ci / Admin1234!", "Employé SOTRA → /mon-espace", ""),
        (35, "admin@cabinet-expertise.ci / Admin1234!", "Admin Cabinet Expertise CI → /dashboard", ""),
        (36, "employe2@cabinet-expertise.ci / Admin1234!", "Employé Cabinet → /mon-espace", ""),
        (37, "coulwao@gmail.com / Openlab2025!", "Admin OpenLab Consulting → /dashboard", ""),
        (39, "URL PRODUCTION", "", ""),
        (40, "https://nexusrh.openlabconsulting.com", "Application web NexusRH CI", ""),
        (41, "https://nexusrh.openlabconsulting.com/api/docs", "Swagger API (debug)", ""),
    ]

    section_rows = {3, 11, 17, 21, 29, 39}
    header_rows  = {4, 12}

    for (r, a, b, c_val) in rows:
        ws.row_dimensions[r].height = 18
        if r in section_rows:
            ws.merge_cells(f"A{r}:C{r}")
            cell = ws.cell(row=r, column=1, value=f"  {a}")
            cell.fill = PatternFill("solid", fgColor="E85D04")
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            for ci, val in enumerate([a, b, c_val], 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.border = thin_border()
                if r in header_rows:
                    cell.fill = PatternFill("solid", fgColor="DDE3EF")
                    cell.font = Font(bold=True, size=9, name="Calibri")
                else:
                    cell.fill = PatternFill("solid", fgColor=C_WHITE if r % 2 == 0 else C_BGLIGHT)
                    cell.font = Font(size=9, name="Calibri")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    return ws


def make_tickets(wb):
    ws = wb.create_sheet(title="TICKETS")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="  SUIVI DES TICKETS & ANOMALIES — NexusRH CI")
    c.fill = PatternFill("solid", fgColor="7F1D1D")
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    cols = [
        ("N° Ticket", 14), ("Lié au test", 14), ("Module", 14),
        ("Titre de l'anomalie", 40), ("Description détaillée", 44),
        ("Priorité", 14), ("Sévérité", 14),
        ("Statut ticket", 16), ("Assigné à", 16), ("Date création", 14),
        ("Date résolution", 14), ("Commentaires", 40),
    ]

    sev_colors = {
        "🔴 Critique": "FEE2E2", "🟠 Majeure": "FEF9C3",
        "🟡 Mineure": "DBEAFE", "🟢 Cosmétique": "DCFCE7",
    }

    ws.row_dimensions[2].height = 22
    for ci, (hdr, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width
        c = ws.cell(row=2, column=ci, value=hdr)
        c.fill      = PatternFill("solid", fgColor="7F1D1D")
        c.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()

    # Example ticket rows
    tickets = [
        ("NRH-001", "CNP-006", "CNPS", "RNS PDF retourne 500", "Le fichier rns-template.pdf n'était pas copié dans dist/assets dans le Dockerfile.",
         "P1 — Critique", "🔴 Critique", "✅ Résolu", "DevOps", TODAY, TODAY, "Fix: COPY src/assets → dist/assets dans API Dockerfile. Déployé en prod."),
        ("NRH-002", "AUTH-003", "AUTH", "Login page silencieuse sans afficher l'erreur 401", "L'intercepteur Axios interceptait les routes /auth/ et faisait window.location.href='/login' avant que le composant puisse afficher l'erreur.",
         "P1 — Critique", "🔴 Critique", "✅ Résolu", "Frontend", TODAY, TODAY, "Fix: intercepteur 401 skip /auth/ routes. Commité et déployé."),
        ("NRH-003", "AUTH-011", "AUTH", "must_change_password déclenché sur tous les comptes seedés", "last_login_at = NULL dans le seed → must_change_password = true pour tous les utilisateurs.",
         "P2 — Haute", "🟠 Majeure", "✅ Résolu", "Backend", TODAY, TODAY, "Fix: seed.ts ajoute last_login_at = now() pour tous les users seedés."),
    ]

    status_list_tick = '"✅ Résolu,🔄 En cours,⏳ À faire,🚫 Abandonné"'
    sev_list         = '"🔴 Critique,🟠 Majeure,🟡 Mineure,🟢 Cosmétique"'

    for ri, tkt in enumerate(tickets, 3):
        ws.row_dimensions[ri].height = 55
        for ci, val in enumerate(tkt, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.font      = Font(size=9, name="Calibri")
            c.border    = thin_border()
            bg = C_BGLIGHT if ri % 2 == 0 else C_WHITE
            if ci == 1:
                c.fill = PatternFill("solid", fgColor="7F1D1D")
                c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 6:
                pb, pf = PRIO_COLORS.get(val, ("FFFFFF", "000000"))
                c.fill = PatternFill("solid", fgColor=pb)
                c.font = Font(bold=True, color=pf, size=9, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = PatternFill("solid", fgColor=bg)

        dv1 = DataValidation(type="list", formula1=status_list_tick, allow_blank=True)
        dv1.sqref = f"H{ri}"
        ws.add_data_validation(dv1)
        dv2 = DataValidation(type="list", formula1=sev_list, allow_blank=True)
        dv2.sqref = f"G{ri}"
        ws.add_data_validation(dv2)
        dv3 = DataValidation(type="list", formula1=PRIO_LIST, allow_blank=True)
        dv3.sqref = f"F{ri}"
        ws.add_data_validation(dv3)

    # Empty rows for new tickets
    for ri in range(len(tickets) + 3, len(tickets) + 23):
        ws.row_dimensions[ri].height = 40
        for ci in range(1, len(cols) + 1):
            c = ws.cell(row=ri, column=ci, value="")
            c.border    = thin_border()
            c.fill      = PatternFill("solid", fgColor=C_BGLIGHT if ri % 2 == 0 else C_WHITE)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if ci == 1:
                c.value = f"NRH-{ri - 2:03d}"
                c.fill  = PatternFill("solid", fgColor="FEF2F2")
                c.font  = Font(color="9CA3AF", size=9, name="Calibri")
            dv1 = DataValidation(type="list", formula1=status_list_tick, allow_blank=True)
            dv1.sqref = f"H{ri}"
            ws.add_data_validation(dv1)
            dv2 = DataValidation(type="list", formula1=sev_list, allow_blank=True)
            dv2.sqref = f"G{ri}"
            ws.add_data_validation(dv2)
            dv3 = DataValidation(type="list", formula1=PRIO_LIST, allow_blank=True)
            dv3.sqref = f"F{ri}"
            ws.add_data_validation(dv3)

    return ws


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    logo_path = "OPENLAB.png"

    # Dashboard first (active sheet)
    make_dashboard(wb, logo_path)

    # Guide second
    make_guide(wb)

    # Module sheets
    for module_key, module_data in MODULES.items():
        make_module_sheet(wb, module_key, module_data)

    # Tickets sheet last
    make_tickets(wb)

    out = "NEXUSRH_CI_Test_Plan.xlsx"
    wb.save(out)
    total = sum(len(m["cases"]) for m in MODULES.values())
    try:
        print(f"✅ {out} généré — {len(MODULES)} modules — {total} scénarios")
    except UnicodeEncodeError:
        print(f"[OK] {out} genere - {len(MODULES)} modules - {total} scenarios")


main()
